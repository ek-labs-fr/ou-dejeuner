"""Export data/restaurants.json to data/restaurants.xlsx for human inspection.

Adds a computed walkMin column (haversine x 1.3 / 80 m/min from the office
anchor) so the sheet is sortable by walking time without further processing.
"""
from __future__ import annotations

import json
import math
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

sys.path.insert(0, str(Path(__file__).resolve().parent))
from places import humanize_type, type_icon

# Office anchor — loaded from env so the public repo doesn't pin coords.
load_dotenv(Path(__file__).resolve().parent.parent / ".env")


def _read_coord(name: str) -> float:
    raw = os.environ.get(name)
    if not raw:
        raise SystemExit(f"{name} env var is required (define in .env)")
    try:
        return float(raw)
    except ValueError:
        raise SystemExit(f"{name} must be a number, got: {raw!r}")


OFFICE_LAT = _read_coord("OFFICE_LAT")
OFFICE_LNG = _read_coord("OFFICE_LNG")

# Walking-time formula matches product-requirements-v2.md:140.
EARTH_RADIUS_M = 6_371_000
WALKING_SPEED_M_PER_MIN = 80.0
DETOUR_FACTOR = 1.3


def _haversine_meters(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    d_lat = math.radians(lat2 - lat1)
    d_lng = math.radians(lng2 - lng1)
    a = (
        math.sin(d_lat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(d_lng / 2) ** 2
    )
    return 2 * EARTH_RADIUS_M * math.asin(math.sqrt(a))


def _walk_minutes(lat: float | None, lng: float | None) -> int | None:
    if lat is None or lng is None:
        return None
    distance = _haversine_meters(OFFICE_LAT, OFFICE_LNG, lat, lng)
    return round(distance * DETOUR_FACTOR / WALKING_SPEED_M_PER_MIN)


# Mirrors lib/walking.ts:compassFromOffice so the xlsx and the app agree.
_COMPASS_8 = (
    "North", "North-East", "East", "South-East",
    "South", "South-West", "West", "North-West",
)


def _compass_from_office(lat: float | None, lng: float | None) -> str | None:
    if lat is None or lng is None:
        return None
    phi1 = math.radians(OFFICE_LAT)
    phi2 = math.radians(lat)
    d_lng = math.radians(lng - OFFICE_LNG)
    y = math.sin(d_lng) * math.cos(phi2)
    x = math.cos(phi1) * math.sin(phi2) - math.sin(phi1) * math.cos(phi2) * math.cos(d_lng)
    deg = (math.degrees(math.atan2(y, x)) + 360) % 360
    return _COMPASS_8[round(deg / 45) % 8]


def main() -> int:
    project_root = Path(__file__).resolve().parent.parent
    in_path = project_root / "data" / "restaurants.json"
    out_path = project_root / "data" / "restaurants.xlsx"

    if not in_path.exists():
        print(f"Input file not found: {in_path}")
        return 1

    data = json.loads(in_path.read_text(encoding="utf-8"))
    places = data["places"]

    wb = Workbook()
    ws = wb.active
    ws.title = "restaurants"

    headers = [
        "name",                # A
        "primaryTypeIcon",     # B  emoji
        "primaryTypeLabel",    # C  human-friendly
        "primaryType",         # D  raw Google token
        "walkMin",             # E
        "direction",           # F  cardinal bearing from office
        "address",             # G
        "priceLevel",          # H
        "businessStatus",      # I
        "allTypesLabels",      # J  human-friendly w/ icons inline
        "allTypes",            # K  raw
        "openingHours",        # L
        "googleMapsUri",       # M
        "latitude",            # N
        "longitude",           # O
        "placeId",             # P
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for p in places:
        loc = p.get("location") or {}
        lat = loc.get("latitude")
        lng = loc.get("longitude")
        opening = p.get("regularOpeningHours") or {}
        weekdays = opening.get("weekdayDescriptions") or []
        # Prefer persisted enrichment fields; fall back to on-the-fly compute
        # if a row predates `enrich_labels.py`.
        primary_type = p.get("primaryType")
        primary_label = p.get("primaryTypeLabel") or humanize_type(primary_type)
        primary_icon = p.get("primaryTypeIcon") or type_icon(primary_type)
        types_raw = p.get("types") or []
        types_labels = p.get("typesLabels") or [humanize_type(t) for t in types_raw]
        types_icons = p.get("typesIcons") or [type_icon(t) for t in types_raw]
        # Inline the icon next to each label in the all-types column so the
        # full list reads as "🇫🇷 French; 🍽️ Restaurant; 🍽️ Food; ..."
        types_with_icons = [
            f"{icon} {label}".strip()
            for icon, label in zip(types_icons, types_labels)
        ]

        ws.append([
            (p.get("displayName") or {}).get("text", ""),
            primary_icon,
            primary_label,
            primary_type,
            _walk_minutes(lat, lng),
            _compass_from_office(lat, lng),
            p.get("formattedAddress"),
            p.get("priceLevel"),
            p.get("businessStatus"),
            "; ".join(types_with_icons),
            "; ".join(types_raw),
            "\n".join(weekdays),
            p.get("googleMapsUri"),
            lat,
            lng,
            p.get("id"),
        ])

    # Freeze the header, enable column auto-filter, set sensible widths.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    column_widths = {
        "A": 36, "B": 6,  "C": 24, "D": 22, "E": 9,  "F": 12, "G": 52, "H": 14,
        "I": 14, "J": 60, "K": 50, "L": 32, "M": 50, "N": 11, "O": 11, "P": 32,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Wrap multi-line cells: allTypesLabels (J), allTypes (K), openingHours (L).
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
        row[9].alignment = wrap   # J
        row[10].alignment = wrap  # K
        row[11].alignment = wrap  # L

    # Make googleMapsUri (column M) clickable.
    for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
        cell = row[12]  # M
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    wb.save(out_path)
    print(f"Wrote {len(places)} rows to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
