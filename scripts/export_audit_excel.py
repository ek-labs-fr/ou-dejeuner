"""Export medium- and low-severity audit results to Excel for admin review.

Reads `data/restaurants-audited.json`, writes
`data/restaurants-audit-review.xlsx` with one sheet per severity bucket:
- "Medium" — flagged for review (bars, desserts, delivery, no-lunch-hours, ...)
- "Low"    — informational (no hours, pastry/coffee/cafe primary types)

High-severity places (auto-exclude recommended) are not included; they're
covered by the audit summary itself.

Each row carries an empty "decision" column so the admin can mark Keep /
Hide / Unsure inline before importing the result back to the catalogue.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = [
    ("decision", 14),                # admin fills in: Keep / Hide / Unsure
    ("name", 42),
    ("primaryTypeIcon", 6),          # emoji
    ("primaryTypeLabel", 24),        # human-friendly primary type
    ("primaryType", 22),             # raw Google token (kept for reference)
    ("typesLabels", 60),             # human-friendly all-types list w/ icons inline
    ("tags", 36),                    # audit tags
    ("address", 48),
    ("businessStatus", 18),
    ("googleMapsUri", 28),
    ("placeId", 32),
]

HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)


def _row_for(place: dict) -> list:
    audit = place.get("audit", {})
    icons = place.get("typesIcons", [])
    labels = place.get("typesLabels", [])
    types_with_icons = [
        f"{icon} {label}".strip() for icon, label in zip(icons, labels)
    ]
    return [
        "",  # decision — empty for admin to fill in
        place.get("displayName", {}).get("text", ""),
        place.get("primaryTypeIcon", ""),
        place.get("primaryTypeLabel", ""),
        place.get("primaryType", ""),
        ", ".join(types_with_icons),
        ", ".join(audit.get("tags", [])),
        place.get("formattedAddress", ""),
        place.get("businessStatus", ""),
        place.get("googleMapsUri", ""),
        place.get("id", ""),
    ]


def _write_sheet(ws, places: list[dict]) -> None:
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Sort by primaryType, then name — keeps similar things adjacent for the admin.
    places_sorted = sorted(
        places,
        key=lambda p: (
            p.get("primaryType") or "~",
            (p.get("displayName") or {}).get("text", ""),
        ),
    )
    for p in places_sorted:
        ws.append(_row_for(p))

    # Hyperlink the googleMapsUri cells. Resolve column index from COLUMNS so
    # this stays correct if columns are reordered.
    gmaps_col = next(i for i, (n, _) in enumerate(COLUMNS, start=1) if n == "googleMapsUri")
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=gmaps_col)
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = Font(color="FF0563C1", underline="single")
            cell.value = "Open in Google Maps"

    # Column widths.
    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row + first column (decision).
    ws.freeze_panes = "B2"

    # Wrap long-text columns so the cells stay readable.
    wrap_col_names = {"typesLabels", "tags", "address"}
    wrap_indices = [
        i for i, (n, _) in enumerate(COLUMNS, start=1) if n in wrap_col_names
    ]
    for row in ws.iter_rows(min_row=2, max_col=len(COLUMNS)):
        for col_idx in wrap_indices:
            row[col_idx - 1].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row[0].row].height = 32

    # Wrap as a real Excel Table (filterable, banded rows).
    last_col = get_column_letter(len(COLUMNS))
    table = Table(
        displayName=f"audit_{ws.title.lower()}",
        ref=f"A1:{last_col}{ws.max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
    )
    ws.add_table(table)

    # Decision column dropdown: Keep / Hide / Unsure.
    dv = DataValidation(
        type="list",
        formula1='"Keep,Hide,Unsure"',
        allow_blank=True,
        showDropDown=False,  # False means the dropdown IS shown (Excel quirk)
    )
    dv.add(f"A2:A{ws.max_row}")
    ws.add_data_validation(dv)


def main() -> int:
    project_root = Path(__file__).resolve().parent.parent
    src = project_root / "data" / "restaurants-audited.json"
    dst = project_root / "data" / "restaurants-audit-review.xlsx"

    if not src.exists():
        print(f"missing {src} — run scripts/audit_places.py first", file=sys.stderr)
        return 1

    payload = json.loads(src.read_text(encoding="utf-8"))
    places = payload.get("places", [])

    medium = [p for p in places if p.get("audit", {}).get("severity") == "medium"]
    low = [p for p in places if p.get("audit", {}).get("severity") == "low"]

    wb = Workbook()
    ws_medium = wb.active
    ws_medium.title = "Medium"
    _write_sheet(ws_medium, medium)

    ws_low = wb.create_sheet("Low")
    _write_sheet(ws_low, low)

    wb.save(dst)
    print(f"Wrote {dst.relative_to(project_root)}")
    print(f"  Medium: {len(medium)} rows")
    print(f"  Low:    {len(low)} rows")
    return 0


if __name__ == "__main__":
    sys.exit(main())
